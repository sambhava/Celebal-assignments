"""R12 remediation: OOXML XML hardening + backend pinning.

Three controls:

1. ``screen_ooxml_bytes`` -- reject any XML member carrying a DTD (``<!DOCTYPE``)
   or an entity declaration (``<!ENTITY``) before it reaches the parser. OOXML
   produced by real spreadsheet software never needs a DTD, so this is a clean
   fail-closed screen that kills billion-laughs and external-entity XXE at the
   door regardless of backend.
2. ``assert_no_resolve_entities_kwarg`` -- regression guard proving the spec's
   fictional ``resolve_entities`` kwarg is not (silently) relied upon: openpyxl
   has no such parameter.
3. ``assert_xml_backend_pinned`` -- assert openpyxl uses the expected XML
   backend so entity-handling posture is deterministic across reviewer machines
   (lxml present vs absent must not change security).

The ``--network none`` sandbox and ``PARSE_RSS_BYTES`` remain the ultimate
backstops; these controls make the parser itself refuse the attack.
"""

from __future__ import annotations

import inspect
import re

from copilot.errors import SecurityViolation

_DTD = re.compile(rb"<!DOCTYPE", re.IGNORECASE)
_ENTITY = re.compile(rb"<!ENTITY", re.IGNORECASE)

#: The XML backend we pin to. Stdlib ElementTree does not fetch external
#: entities; combined with the DTD screen this is the deterministic baseline.
EXPECTED_BACKEND_LXML = False


def screen_ooxml_bytes(data: bytes) -> None:
    """Raise ``SecurityViolation`` if the XML declares a DTD or entities."""
    if _DTD.search(data):
        raise SecurityViolation("OOXML XML contains a DTD (<!DOCTYPE); rejected (R12, XXE/billion-laughs defense).")
    if _ENTITY.search(data):
        raise SecurityViolation("OOXML XML declares an entity (<!ENTITY); rejected (R12).")


def assert_no_resolve_entities_kwarg() -> None:
    """Fail if code assumes openpyxl.load_workbook accepts ``resolve_entities``."""
    import openpyxl

    params = inspect.signature(openpyxl.load_workbook).parameters
    if "resolve_entities" in params:
        raise AssertionError("openpyxl.load_workbook now has resolve_entities; revisit R12.")


def assert_xml_backend_pinned() -> None:
    """Assert openpyxl's XML backend matches the pinned expectation."""
    try:
        from openpyxl.xml.functions import LXML
    except Exception as exc:  # pragma: no cover - defensive
        raise AssertionError("cannot determine openpyxl XML backend: {}".format(exc))
    if bool(LXML) != EXPECTED_BACKEND_LXML:
        raise AssertionError(
            "openpyxl XML backend LXML={} but pinned expectation is {}; "
            "entity-handling posture would differ -- pin the backend.".format(
                bool(LXML), EXPECTED_BACKEND_LXML
            )
        )
